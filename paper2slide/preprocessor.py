import re
from gingerit.gingerit import GingerIt
import openpyxl
import nltk

# Download the necessary resources for sentence tokenization
nltk.download('punkt')


def get_cleaned_text(text):
    text = remove_et_al_period(text)
    text = remove_hyphen_spaces(text)
    text = correct_grammar(text)
    return text

def remove_et_al_period(text):
    # "et al."을 "id est alii"로 변경합니다.
    text = re.sub(r'et al\.', 'et alii', text)
    
    # "e.g."를 "exempli gratia"로 변경합니다.
    text = re.sub(r'e\.g\.', 'exempli gratia', text)
    text = re.sub(r'\.\.\.', '~~~', text)
    # "i.e."를 "id est"로 변경합니다.
    text = re.sub(r'i\.e\.', 'id est', text)
    
    # "cf."를 "confer"로 변경합니다.
    text = re.sub(r'cf\.', 'confer', text)
    
    # 숫자의 소수 형태(ex: 4.5)를 찾아 쉼표로 변경합니다.
    text = re.sub(r'(\d+)\.(\d+)', r'\1,\2', text)
    
    return text

def replace_et_al(text):
    # "et alii"을 "et al."로 변경합니다.
    text = re.sub(r'et Alii', 'et al.', text)
    
    # "exempli gratia"를 "e.g."로 변경합니다.
    text = re.sub(r'exempli gratia', 'e.g.', text)
    
    # "id est"를 "i.e."로 변경합니다.
    text = re.sub(r'id EST', 'i.e.', text)
    
    # "confer"를 "cf."로 변경합니다.
    text = re.sub(r'confer', 'cf.', text)
    text = re.sub(r'~~~', '...', text)
    
    return text

def remove_hyphen_spaces(text):
    # -로 이어진 단어들의 패턴을 정의
    pattern = re.compile(r'(\w+)-\s+(\w+)')
    # 패턴과 일치하는 문자열을 검색하여 공백 제거
    corrected_text = pattern.sub(r'\1\2', text)
    # 결과 반환
    return corrected_text

def correct_grammar(text):
    # 문장을 온점 기준으로 분리
    sentences = text.split('.')
    # 문장의 마지막이 온점이 아니라면, 마지막 문장에서 온점을 추가
    if sentences[-1] != '':
        sentences[-1] += '.'
    # GingerIt 클래스의 인스턴스 생성
    parser = GingerIt()
    # 수정된 문장을 저장할 리스트
    corrected_sentences = []
    # 각 문장에 대해 문법 검사 수행
    for sentence in sentences:
        # 공백 제거
        sentence = sentence.strip()
        # 빈 문자열이라면 다음 문장으로 건너뜀
        if sentence == '':
            continue
        # 문법 오류 수정
        result = parser.parse(sentence)
        corrected_sentence = result['result']
        corrected_sentences.append(corrected_sentence)
    # 수정된 문장을 온점으로 이어붙임
    result_text = '. '.join(corrected_sentences)
    # 결과 반환
    return result_text


def extract_references(text):
    # Remove any leading or trailing whitespaces
    text = text.strip()
    
    # Split the text into individual references using regular expressions
    references = re.split(r'\[\d+\]', text)
    
    # Remove any empty references
    references = [ref.strip() for ref in references if ref.strip()]
    
    # Create a dictionary to store the references with their corresponding numbers
    references_dict = {}
    for i, ref in enumerate(references, start=1):
        references_dict[i] = ref
    
    return references_dict


def convert_references_section_title(text):
    # 패턴 매칭을 사용하여 "4. References"와 같은 형태를 "REFERENCES"로 변환
    converted_text = re.sub(r'^\d*\.*\s*References', 'REFERENCES', text, flags=re.IGNORECASE)
    
    return converted_text


def extract_table(data):
    paths = []
    for i in range(len(data)):
        # Extract table dimensions
        row_count = data[i]['row_count']
        column_count = data[i]['column_count']

        # Create empty workbook and active sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Fill table with cell content
        for cell in data[i]['cells']:
            row_index = cell['row_index']
            column_index = cell['column_index']
            content = cell['content']

            # Check if cell has columnSpan or rowSpan
            if 'column_span' in cell:
                column_span = cell['column_span']
                end_column = column_index + column_span - 1
                sheet.cell(row=row_index + 1, column=column_index + 1, value=content)
                sheet.merge_cells(start_row=row_index + 1, start_column=column_index + 1, end_row=row_index + 1,
                                end_column=end_column + 1)
            elif 'row_span' in cell:
                row_span = cell['row_span']
                end_row = row_index + row_span - 1
                sheet.cell(row=row_index + 1, column=column_index + 1, value=content)
                sheet.merge_cells(start_row=row_index + 1, start_column=column_index + 1, end_row=end_row + 1,
                                end_column=column_index + 1)
            else:
                sheet.cell(row=row_index + 1, column=column_index + 1, value=content)

        # Save workbook as XLSX file
        path = 'table_' + str(i+1) + '.xlsx'
        workbook.save(path)
        paths.append(path)
    return paths

def split_sentences(text):
    # Tokenize the paragraph into sentences
    sentences = nltk.sent_tokenize(text)

    return sentences

def data_reconstruction(data):
    result = []
    current_section_heading = None
    previous_section_heading = None

    for sentence in data["sentences"]:
        if sentence["role"] == "sectionHeading":
            if current_section_heading is not None:
                result.append({
                    "title" : previous_section_heading,
                    "content": current_section_content,
                    "summarized": current_section_summarized,
                    # "tables": current_section_tables,
                    # "figures": current_section_figures
                })

            previous_section_heading = sentence["content"]
            current_section_heading = sentence["content"]
            current_section_content = None
            current_section_summarized = None
            current_section_tables = None
            current_section_figures = None
        elif sentence["role"] is None and current_section_heading is not None:
            current_section_content = sentence["content"]
            current_section_summarized = sentence["summarized"]
            # current_section_tables = sentence["tables"]
            # current_section_figures = sentence["figures"]

    # Append the last section
    if current_section_heading is not None:
        result.append({
            "title": current_section_heading,
            "content": current_section_content,
            "summarized": current_section_summarized,
            # "tables": current_section_tables,
            # "figures": current_section_figures
        })

    return result
def find_pattern_match_position(content, summarized, find):
    if find.lower() == "figure":
        pattern = r"(?i)(Fig(?:\.|ure)?s?\s*\d+(?:\s*and\s*\d+)?)(?!:)"
    elif find.lower() == "table":
        pattern = r"(?i)(Table(?:s)?\s*\d+(?:\s*and\s*\d+)?)(?!:)"
    else:
        pattern = r"(?i)(" + re.escape(find) + r"\s+\d+)(?!:)"

    pattern_matches = re.finditer(pattern, content)
    positions = []
    for match in pattern_matches:
        match_start = match.start()
        for i in range(-1, len(summarized) - 1):
            if i == -1:
                start_index = 0
            else:
                start_index = content.index(summarized[i][4:-4]) + len(summarized[i][4:-4])
            end_index = content.index(summarized[i + 1][4:-4])
            if start_index <= match_start < end_index:
                group = match.group(0).lower().replace(" ", "_")
                if "figure" in group or "fig." in group:
                    group = group.replace("fig.", "").replace("figures", "").replace("figure", "").strip("_s")
                    if "and" in group:
                        split_group = group.split("and")
                        for item in split_group:
                            positions.append(("figure_" + item.strip().strip("_"), i + 1))
                    else:
                        positions.append(("figure_" + group.strip().strip("_"), i + 1))
                elif "table" in group:
                    group = group.replace("tables", "").replace("table", "").strip("_s")
                    if "and" in group:
                        split_group = group.split("and")
                        for item in split_group:
                            positions.append(("table_" + item.strip().strip("_"), i + 1))
                    else:
                        positions.append(("table_" + group.strip().strip("_"), i + 1))
                else:
                    positions.append((group.strip().strip("_"), i + 1))
        if positions == [] and match:
            group = match.group(0).lower().replace(" ", "_")
            if "figure" in group or "fig." in group:
                group = group.replace("fig.", "").replace("figures", "").replace("figure", "").strip("_s")
                if "and" in group:
                    split_group = group.split("and")
                    for item in split_group:
                        positions.append(("figure_" + item.strip().strip("_"), i + 1))
                else:
                    positions.append(("figure_" + group.strip().strip("_"), i + 1))
            elif "table" in group:
                group = group.replace("tables", "").replace("table", "").strip("_s")
                if "and" in group:
                    split_group = group.split("and")
                    for item in split_group:
                        positions.append(("table_" + item.strip().strip("_"), i + 1))
                else:
                    positions.append(("table_" + group.strip().strip("_"), i + 1))
            else:
                positions.append((group.strip().strip("_"), i + 1))
                
    return positions